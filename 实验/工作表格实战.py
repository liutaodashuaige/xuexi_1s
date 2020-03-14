import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def update_money(filename):
    ex = xl.load_workbook(filename)
    sheet = ex['Sheet1']
    for hezi in range(2, sheet.max_row + 1):
        cell = sheet.cell(hezi, 3)
        update_cell_value = cell.value * 2
        update_cell = sheet.cell(hezi, 4)
        update_cell.value = update_cell_value
        print(update_cell.value)
    my_data = Reference(sheet,
                        min_row=2,
                        max_row=sheet.max_row,
                        min_col=4,
                        max_col=4)
    chart = BarChart()
    chart.add_data(my_data)
    sheet.add_chart(chart, 'e1')
    ex.save('python_demo1.1.xlsx')