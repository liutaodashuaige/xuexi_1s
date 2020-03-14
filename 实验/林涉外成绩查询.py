# encoding:utf-8
from selenium import webdriver
from selenium.webdriver.support.select import Select
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import wcwidth
import time
#函数
def 查询成绩制表(driver,学期数):
    # 存为excel表
    wb = Workbook()
    for i in range(2, 学期数):
        # 切换iframe,进行查找
        driver.switch_to.frame("cjcx_query_frm")
        select_time = driver.find_element_by_name('kksj')
        select_time.click()
        time.sleep(1)
        Select(select_time).select_by_index(i)  # 通过索引选择筛选条件
        driver.find_element_by_id('btn_query').click()
        # 还原iframe位置
        driver.switch_to.default_content()
        time.sleep(1)

        # 切换iframe，读取表格
        # 读表头
        driver.switch_to.frame("cjcx_list_frm")
        lstheader = []
        table_top_list = driver.find_element_by_xpath("//*[@id='dataList']/tbody/tr").find_elements_by_tag_name('th')
        # 存表头
        for c, top in enumerate(table_top_list):
            lstheader.append(top.text)
        # 读每一行内容
        lstcontent = []
        table_tr_list = driver.find_element_by_xpath("//*[@id='dataList']/tbody").find_elements_by_tag_name('tr')
        # 存内容
        for r, tr in enumerate(table_tr_list):
            if r == 0:
                continue
            lst_row = []
            table_td_list = tr.find_elements_by_tag_name('td')
            for c, td in enumerate(table_td_list):
                lst_row.append(td.text)
            lstcontent.append(lst_row)

        # 写入工作表
        wb.create_sheet(lstcontent[1][1])
        sheet = wb[lstcontent[1][1]]
        sheet.append(lstheader)
        for tr in lstcontent:
            sheet.append(tr)
        #表格美化
        tianchon_fill = PatternFill(fill_type='solid', fgColor="FFC125")
        duiqi_alignment = Alignment(horizontal='center', vertical='center')

        for hang in range(sheet.min_row,sheet.max_row+1):
            for lie in range(sheet.min_column,sheet.max_column+1):
                sheet.cell(hang, lie).alignment = duiqi_alignment
                if lie in [4,5,6]:
                    sheet.cell(hang, lie).fill = tianchon_fill
                if lie in [2,3,4,11]:
                    sheet.column_dimensions[get_column_letter(lie)].width = 16


        print(f"{lstcontent[1][1]},{'成绩单已保存'}")

        # 还原iframe位置
        driver.switch_to.default_content()
    # 存表
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(r'成绩表.xlsx')
    driver.close()
#函数
def 模拟登录(账号,密码):
    #模拟登陆
    chrome_driver = r"C:\Program Files (x86)\Google\Chrome\Application\chromedriver"
    driver = webdriver.Chrome(chrome_driver)
    driver.set_window_position(20,20)
    driver.set_window_size(700,1000)
    driver.get("http://54.222.196.251:81/znlykjdxswxy_jsxsd/")
    driver.find_element_by_id('userAccount').clear()
    driver.find_element_by_id('userAccount').send_keys(账号)
    driver.find_element_by_id('userPassword').clear()
    driver.find_element_by_id('userPassword').send_keys(密码)
    driver.find_element_by_id('btnSubmit').click()

    time.sleep(1)

    driver.find_element_by_xpath("/html/body/div[2]/div[1]/div[1]/ul/li[3]/a").click()
    driver.find_element_by_xpath("/html/body/div[3]/div[1]/ul[2]/li/a").click()
    return driver

if __name__ == '__main__':
    查询成绩制表(模拟登录('20177587','qwer1234.'),7)

